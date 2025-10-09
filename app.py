# app.py
# ============================================================
# Análisis MICMAC Interactivo - Implementación Académica
# by Martín Pratto
# Versión 3.5 - Con optimización automática de parámetros
# ============================================================
"""
Implementación open-source del algoritmo MICMAC (Matriz de Impactos 
Cruzados - Multiplicación Aplicada a una Clasificación) según la 
metodología de Michel Godet (1990).

Referencias:
- Godet, M. (1990). From Anticipation to Action: A Handbook of 
  Strategic Prospective. UNESCO Publishing.
- Godet, M., & Durance, P. (2011). Strategic Foresight for 
  Corporate and Regional Development.
"""

import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import io
from datetime import datetime
from openpyxl import load_workbook
from collections import Counter

# Configuración de matplotlib
plt.rcParams.update({
    "axes.titlesize": 18,
    "axes.labelsize": 14,
    "xtick.labelsize": 10,
    "ytick.labelsize": 10,
    "figure.dpi": 100
})

# ============================================================
# CONFIGURACIÓN DE PÁGINA
# ============================================================
st.set_page_config(
    page_title="Análisis MICMAC Interactivo",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================
# ENCABEZADO
# ============================================================
st.markdown("""
# 📊 Análisis MICMAC Interactivo  
### Análisis Estructural de Sistemas Complejos
**by Martín Pratto** • *Versión 3.5 - Con Optimización Automática*

---
""")

with st.expander("ℹ️ Acerca de esta herramienta", expanded=False):
    st.markdown("""
    ### Metodología MICMAC
    
    El método MICMAC (Matriz de Impactos Cruzados - Multiplicación Aplicada a una Clasificación) 
    es una técnica de análisis estructural desarrollada por **Michel Godet** en el contexto de la 
    prospectiva estratégica francesa.
    
    **¿Qué hace esta herramienta?**
    - Analiza sistemas complejos identificando variables clave
    - Calcula influencias **directas** (matriz original) e **indirectas** (propagación)
    - Clasifica variables en 4 cuadrantes estratégicos
    - Genera rankings, gráficos y reportes ejecutivos
    - **NUEVO:** Optimización automática de parámetros α y K
    
    **Cómo usar:**
    1. **Sube tu matriz Excel** (variables como filas/columnas)
    2. **Elige modo automático** (recomendado) o manual
    3. **Explora resultados** interactivos y descarga reportes
    """)

with st.expander("📚 Referencias Bibliográficas", expanded=False):
    st.markdown("""
    - **Godet, M. (1990).** *From Anticipation to Action: A Handbook of Strategic Prospective.* UNESCO Publishing.
    - **Godet, M., & Durance, P. (2011).** *Strategic Foresight for Corporate and Regional Development.* 
      Fondation Prospective et Innovation, UNESCO.
    - **Arcade, J., Godet, M., Meunier, F., & Roubelat, F. (2004).** *Structural analysis with the MICMAC method.* 
      Futures Research Methodology, AC/UNU Millennium Project.
    """)

# ============================================================
# FUNCIONES CORE MICMAC
# ============================================================

def ensure_square_from_df(df: pd.DataFrame) -> pd.DataFrame:
    """Convierte DataFrame en matriz cuadrada."""
    # Conversión a numérico
    df = df.apply(pd.to_numeric, errors='coerce').fillna(0.0)
    
    # Asegurar que índice y columnas son strings
    df.index = df.index.astype(str)
    df.columns = df.columns.astype(str)
    
    # Intersección
    common = df.index.intersection(df.columns)
    
    if len(common) < 3:
        st.error(f"❌ Solo {len(common)} variables coincidentes. Se necesitan al menos 3.")
        st.write("**Filas:**", df.index.tolist()[:10])
        st.write("**Columnas:**", df.columns.tolist()[:10])
        raise ValueError(f"Insuficientes variables coincidentes: {len(common)}")
    
    # Filtrar
    df = df.loc[common, common].copy()
    
    # Diagonal a 0
    np.fill_diagonal(df.values, 0.0)
    
    # Verificar ceros
    filas_cero = df.index[df.sum(axis=1) == 0].tolist()
    cols_cero = df.columns[df.sum(axis=0) == 0].tolist()
    
    if filas_cero:
        st.warning(f"⚠️ {len(filas_cero)} variables con motricidad = 0")
    
    if cols_cero:
        st.warning(f"⚠️ {len(cols_cero)} variables con dependencia = 0")
    
    return df


def micmac_total(M: np.ndarray, alpha: float, K: int) -> np.ndarray:
    """Calcula la matriz total MICMAC con propagación."""
    M = M.astype(float)
    M_total = M.copy()
    M_power = M.copy()
    
    for k in range(2, K + 1):
        M_power = M_power @ M
        M_total += (alpha ** (k - 1)) * M_power
    
    np.fill_diagonal(M_total, 0.0)
    return M_total


def first_stable_K(M: np.ndarray, alpha: float, K_values=range(2, 15)) -> int:
    """Encuentra el primer K donde el ranking se estabiliza."""
    prev_order = None
    for K in K_values:
        M_tot = micmac_total(M, alpha, K)
        motricidad = M_tot.sum(axis=1)
        order = tuple(np.argsort(-motricidad))
        if prev_order is not None and order == prev_order:
            return K
        prev_order = order
    return max(K_values)


def find_optimal_parameters(M: np.ndarray, max_inflation=50):
    """
    Encuentra parámetros α y K óptimos balanceando convergencia e interpretabilidad.
    """
    alpha_values = [0.9, 0.8, 0.7, 0.6, 0.5, 0.4, 0.3, 0.2]
    K_values = range(2, 10)
    
    valid_configs = []
    
    for alpha in alpha_values:
        for K in K_values:
            M_tot = micmac_total(M, alpha, K)
            
            mot_dir = M.sum(axis=1)
            mot_tot = M_tot.sum(axis=1)
            mot_ind = mot_tot - mot_dir
            
            inflation_ratios = []
            for i in range(len(mot_dir)):
                if mot_dir[i] > 0:
                    inflation_ratios.append(mot_ind[i] / mot_dir[i])
            
            if len(inflation_ratios) > 0:
                avg_inflation = np.mean(inflation_ratios)
                max_value = mot_tot.max()
            else:
                avg_inflation = 0
                max_value = 0
            
            ranking_actual = tuple(np.argsort(-mot_tot))
            
            if K < 9:
                M_tot_next = micmac_total(M, alpha, K+1)
                mot_tot_next = M_tot_next.sum(axis=1)
                ranking_next = tuple(np.argsort(-mot_tot_next))
                is_stable = (ranking_actual == ranking_next)
            else:
                is_stable = True
            
            if avg_inflation <= max_inflation and max_value < 1e6:
                valid_configs.append({
                    'alpha': alpha,
                    'K': K,
                    'inflation': avg_inflation,
                    'max_value': max_value,
                    'stable': is_stable,
                    'score': alpha * 10 - K + (10 if is_stable else 0) - avg_inflation/10
                })
    
    if len(valid_configs) == 0:
        return {
            'alpha': 0.3,
            'K': 2,
            'inflation': 0,
            'stable': False,
            'warning': 'No se encontraron parámetros óptimos, usando valores conservadores'
        }
    
    valid_configs.sort(key=lambda x: x['score'], reverse=True)
    return valid_configs[0]


def validate_micmac_results(M: np.ndarray, M_tot: np.ndarray, alpha: float, K: int):
    """Valida que los resultados sean interpretables."""
    warnings = []
    recommendations = []
    
    mot_dir = M.sum(axis=1)
    mot_tot = M_tot.sum(axis=1)
    mot_ind = mot_tot - mot_dir
    
    inflation_ratios = []
    for i in range(len(mot_dir)):
        if mot_dir[i] > 0:
            inflation_ratios.append(mot_ind[i] / mot_dir[i])
    
    if len(inflation_ratios) > 0:
        avg_inflation = np.mean(inflation_ratios)
        
        if avg_inflation > 100:
            warnings.append(f"⚠️ Inflación promedio muy alta: {avg_inflation:.0f}x")
            recommendations.append("Reducir K a 2-3 o α a 0.3-0.4")
        elif avg_inflation > 50:
            warnings.append(f"⚠️ Inflación moderada-alta: {avg_inflation:.0f}x")
            recommendations.append("Considerar reducir K o α")
        elif avg_inflation > 20:
            warnings.append(f"✓ Inflación aceptable: {avg_inflation:.1f}x")
        else:
            warnings.append(f"✅ Inflación baja: {avg_inflation:.1f}x")
    else:
        avg_inflation = 0
    
    max_value = mot_tot.max()
    
    if max_value > 1e6:
        warnings.append(f"⚠️ Valores muy grandes: {max_value:,.0f}")
        recommendations.append("Los valores son correctos pero difíciles de interpretar")
    elif max_value > 1e4:
        warnings.append(f"✓ Valores en miles: {max_value:,.0f}")
    else:
        warnings.append(f"✅ Valores interpretables: {max_value:.0f}")
    
    return {
        'warnings': warnings,
        'recommendations': recommendations,
        'avg_inflation': avg_inflation,
        'max_value': max_value,
        'is_valid': len([w for w in warnings if '⚠️' in w]) == 0
    }


def classify_quadrant(motricidad, dependencia, mot_threshold, dep_threshold):
    """Clasifica variable según cuadrante MICMAC."""
    if motricidad >= mot_threshold and dependencia < dep_threshold:
        return 'Determinantes'
    elif motricidad >= mot_threshold and dependencia >= dep_threshold:
        return 'Crítico/inestable'
    elif motricidad < mot_threshold and dependencia >= dep_threshold:
        return 'Variables resultado'
    else:
        return 'Autónomas'


def analyze_stability(M: np.ndarray, alpha_values, K_values):
    """Analiza estabilidad del ranking."""
    results = []
    for alpha in alpha_values:
        for K in K_values:
            M_tot = micmac_total(M, alpha, K)
            motricidad = M_tot.sum(axis=1)
            ranking_indices = np.argsort(-motricidad)[:5]
            results.append({
                'alpha': alpha,
                'K': K,
                'top_1': ranking_indices[0],
                'top_2': ranking_indices[1],
                'top_3': ranking_indices[2],
                'top_4': ranking_indices[3],
                'top_5': ranking_indices[4],
                'top_5_str': str(ranking_indices[:5].tolist())
            })
    return pd.DataFrame(results)


# ============================================================
# CARGA DE ARCHIVO
# ============================================================
st.markdown("### 📁 Paso 1: Carga tu Matriz MICMAC")

uploaded_file = st.file_uploader(
    "Sube tu archivo Excel con la matriz de influencias directas:",
    type=["xlsx"],
    help="El archivo debe contener una matriz cuadrada con nombres de variables."
)

if not uploaded_file:
    st.info("👆 Por favor, sube un archivo Excel para comenzar el análisis.")
    
    with st.expander("💡 Formato de archivo esperado"):
        st.markdown("""
        **Estructura del archivo Excel:**
        
        | Variable | Var1 | Var2 | Var3 |
        |----------|------|------|------|
        | Var1     | 0    | 3    | 1    |
        | Var2     | 2    | 0    | 2    |
        | Var3     | 1    | 1    | 0    |
        
        - Primera columna: nombres de variables
        - Valores: intensidad de influencia (0-3 o 0-4)
        - La diagonal será automáticamente 0
        """)
    st.stop()

# ============================================================
# PROCESAMIENTO DEL ARCHIVO
# ============================================================
try:
    wb = load_workbook(uploaded_file, data_only=True)
    sheets = wb.sheetnames
    
    sheet = st.selectbox(
        "Selecciona la hoja con la matriz:",
        options=sheets,
        index=0
    )
    
    uploaded_file.seek(0)
    df_raw = pd.read_excel(uploaded_file, sheet_name=sheet, index_col=0)
    
    # Limpiar columnas
    columnas_a_eliminar = ['SUMA', 'Suma', 'suma', 'Total', 'TOTAL', 'total']
    for col in columnas_a_eliminar:
        if col in df_raw.columns:
            df_raw = df_raw.drop(columns=[col])
    
    # Eliminar filas vacías
    df_raw = df_raw.dropna(how='all')
    
    # Convertir a string
    df_raw.index = df_raw.index.map(lambda x: str(x) if pd.notna(x) else '')
    df_raw.columns = df_raw.columns.map(lambda x: str(x) if pd.notna(x) else '')
    
    # Filtrar vacíos
    df_raw = df_raw.loc[df_raw.index != '', df_raw.columns != '']
    
    df = ensure_square_from_df(df_raw)
    nombres = df.index.tolist()
    M = df.values.astype(float)
    
    st.success(f"✅ Archivo cargado. Hoja: **{sheet}** • Variables: **{len(nombres)}**")
    
    with st.expander("👁️ Vista previa de la matriz"):
        st.dataframe(df.head(10), use_container_width=True)

except Exception as e:
    st.error(f"❌ Error: {str(e)}")
    st.stop()

# ============================================================
# CONFIGURACIÓN DE PARÁMETROS
# ============================================================
st.markdown("### ⚙️ Paso 2: Configura los Parámetros")

modo = st.radio(
    "Modo de configuración:",
    options=['🤖 Automático (Recomendado)', '⚙️ Manual (Avanzado)'],
    index=0
)

if modo == '🤖 Automático (Recomendado)':
    st.info("🔍 Calculando parámetros óptimos...")
    
    with st.spinner("Analizando..."):
        optimal_params = find_optimal_parameters(M, max_inflation=50)
    
    if 'warning' in optimal_params:
        st.warning(optimal_params['warning'])
    
    alpha = optimal_params['alpha']
    K_max = optimal_params['K']
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("α óptimo", f"{alpha}")
    
    with col2:
        st.metric("K óptimo", f"{K_max}")
    
    with col3:
        if optimal_params.get('stable', False):
            st.success("✅ Convergente")
        else:
            st.warning("⚠️ Parcialmente estable")
    
    st.info(f"""
    **Parámetros seleccionados:**
    - α = {alpha}
    - K = {K_max}
    - Inflación estimada: {optimal_params['inflation']:.1f}x
    """)

else:  # Modo Manual
    col1, col2 = st.columns(2)
    
    with col1:
        alpha = st.slider("α (Factor de atenuación)", 0.1, 1.0, 0.5, 0.05)
    
    with col2:
        K_max = st.slider("K (Profundidad)", 2, 10, 4)

col_extra1, col_extra2 = st.columns(2)

with col_extra1:
    usar_mediana = st.checkbox("Usar mediana para umbrales", value=False)

with col_extra2:
    max_etiquetas = st.slider(
        "Máx. etiquetas",
        10,
        min(60, len(nombres)),
        min(30, len(nombres)),
        5
    )

# ============================================================
# CÁLCULOS MICMAC
# ============================================================
st.markdown("### 📊 Paso 3: Resultados del Análisis")

with st.spinner("🔄 Procesando..."):
    mot_dir = M.sum(axis=1)
    dep_dir = M.sum(axis=0)
    
    M_tot = micmac_total(M, alpha, K_max)
    mot_tot = M_tot.sum(axis=1)
    dep_tot = M_tot.sum(axis=0)
    
    mot_ind = mot_tot - mot_dir
    dep_ind = dep_tot - dep_dir
    
    df_all = pd.DataFrame({
        "Motricidad_directa": mot_dir,
        "Motricidad_indirecta": mot_ind,
        "Motricidad_total": mot_tot,
        "Dependencia_directa": dep_dir,
        "Dependencia_indirecta": dep_ind,
        "Dependencia_total": dep_tot
    }, index=nombres)
    
    if usar_mediana:
        mot_threshold = np.median(mot_tot)
        dep_threshold = np.median(dep_tot)
    else:
        mot_threshold = np.mean(mot_tot)
        dep_threshold = np.mean(dep_tot)
    
    df_all['Clasificación'] = df_all.apply(
        lambda row: classify_quadrant(
            row['Motricidad_total'],
            row['Dependencia_total'],
            mot_threshold,
            dep_threshold
        ),
        axis=1
    )
    
    order = np.argsort(-mot_tot)
    ranking_vars = [nombres[i] for i in order]
    
    df_rank = pd.DataFrame({
        "Posición": np.arange(1, len(nombres) + 1),
        "Variable": ranking_vars,
        "Motricidad_total": mot_tot[order],
        "Motricidad_directa": mot_dir[order],
        "Motricidad_indirecta": mot_ind[order],
        "Dependencia_total": dep_tot[order],
        "Clasificación": [df_all.loc[var, 'Clasificación'] for var in ranking_vars]
    })

# VALIDACIÓN
validation = validate_micmac_results(M, M_tot, alpha, K_max)

if not validation['is_valid']:
    st.warning("⚠️ **Advertencias:**")
    for warning in validation['warnings']:
        st.write(warning)
    
    if validation['recommendations']:
        st.info("💡 **Recomendaciones:**")
        for rec in validation['recommendations']:
            st.write(f"• {rec}")
else:
    st.success(f"""
    ✅ Análisis completado
    
    - Inflación: {validation['avg_inflation']:.1f}x
    - Máximo: {validation['max_value']:,.0f}
    - Parámetros: α={alpha}, K={K_max}
    """)

# ============================================================
# TABS
# ============================================================
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "📋 Rankings",
    "📈 Subsistemas",
    "🎯 Eje Estratégico",
    "🔬 Estabilidad",
    "📊 Gráficos",
    "📄 Informe"
])

# TAB 1
with tab1:
    st.markdown(f"### 🏆 Ranking (α={alpha}, K={K_max})")
    
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Variables", len(nombres))
    col2.metric("Determinantes", len(df_all[df_all['Clasificación'] == 'Determinantes']))
    col3.metric("Críticas", len(df_all[df_all['Clasificación'] == 'Crítico/inestable']))
    col4.metric("Resultado", len(df_all[df_all['Clasificación'] == 'Variables resultado']))
    
    st.dataframe(
        df_rank.style.background_gradient(subset=['Motricidad_total'], cmap='YlOrRd'),
        use_container_width=True,
        height=400
    )

# TAB 2
with tab2:
    st.markdown("### 📈 Gráfico de Subsistemas")
    
    fig_sub, ax_sub = plt.subplots(figsize=(16, 12))
    
    colors_map = {
        'Determinantes': '#FF4444',
        'Crítico/inestable': '#1166CC',
        'Variables resultado': '#66BBFF',
        'Autónomas': '#FF9944'
    }
    
    colors = [colors_map[df_all.loc[var, 'Clasificación']] for var in nombres]
    sizes = [100 if df_all.loc[var, 'Clasificación'] == 'Crítico/inestable' else 80 for var in nombres]
    
    ax_sub.scatter(dep_tot, mot_tot, c=colors, s=sizes, alpha=0.7, edgecolors='black', linewidth=1.5)
    
    ax_sub.axvline(dep_threshold, color='black', linestyle='--', linewidth=2, alpha=0.6)
    ax_sub.axhline(mot_threshold, color='black', linestyle='--', linewidth=2, alpha=0.6)
    
    max_mot = max(mot_tot)
    max_dep = max(dep_tot)
    
    ax_sub.text(dep_threshold * 0.5, max_mot * 0.9, 'DETERMINANTES',
                fontsize=13, fontweight='bold', ha='center',
                bbox=dict(boxstyle="round", facecolor="red", alpha=0.6), color='white')
    
    ax_sub.text(max_dep * 0.75, max_mot * 0.9, 'CRÍTICO',
                fontsize=13, fontweight='bold', ha='center',
                bbox=dict(boxstyle="round", facecolor="darkblue", alpha=0.6), color='white')
    
    ax_sub.text(dep_threshold * 0.5, mot_threshold * 0.3, 'AUTÓNOMAS',
                fontsize=13, fontweight='bold', ha='center',
                bbox=dict(boxstyle="round", facecolor="orange", alpha=0.6))
    
    ax_sub.text(max_dep * 0.75, mot_threshold * 0.3, 'RESULTADO',
                fontsize=13, fontweight='bold', ha='center',
                bbox=dict(boxstyle="round", facecolor="lightblue", alpha=0.6))
    
    importantes_idx = order[:min(max_etiquetas, len(nombres))]
    for i in importantes_idx:
        ax_sub.annotate(
            nombres[i][:25],
            (dep_tot[i], mot_tot[i]),
            xytext=(5, 5), textcoords='offset points',
            fontsize=8, fontweight='bold',
            bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8),
            arrowprops=dict(arrowstyle='->', color='gray', alpha=0.5)
        )
    
    ax_sub.set_xlabel("Dependencia Total", fontweight='bold', fontsize=14)
    ax_sub.set_ylabel("Motricidad Total", fontweight='bold', fontsize=14)
    ax_sub.set_title(f"SUBSISTEMAS (α={alpha}, K={K_max})", fontweight='bold', fontsize=16)
    ax_sub.grid(True, alpha=0.3)
    
    from matplotlib.lines import Line2D
    legend_elements = [
        Line2D([0], [0], marker='o', color='w', markerfacecolor='#FF4444', markersize=10, label='Determinantes'),
        Line2D([0], [0], marker='o', color='w', markerfacecolor='#1166CC', markersize=10, label='Crítico'),
        Line2D([0], [0], marker='o', color='w', markerfacecolor='#66BBFF', markersize=10, label='Resultado'),
        Line2D([0], [0], marker='o', color='w', markerfacecolor='#FF9944', markersize=10, label='Autónomas')
    ]
    ax_sub.legend(handles=legend_elements, loc='upper left', fontsize=11)
    
    st.pyplot(fig_sub)
    
    img = io.BytesIO()
    fig_sub.savefig(img, format='png', dpi=300, bbox_inches='tight')
    img.seek(0)
    st.download_button("📥 Descargar", img, f"subsistemas_a{alpha}_k{K_max}.png", "image/png")

# TAB 3
with tab3:
    st.markdown("### 🎯 Eje Estratégico")
    
    fig_est, ax_est = plt.subplots(figsize=(14, 11))
    
    max_dep_norm = max(dep_tot) if max(dep_tot) > 0 else 1
    max_mot_norm = max(mot_tot) if max(mot_tot) > 0 else 1
    
    strategic_scores = []
    for i in range(len(nombres)):
        x_norm = dep_tot[i] / max_dep_norm
        y_norm = mot_tot[i] / max_mot_norm
        dist_to_axis = abs(y_norm - x_norm) / np.sqrt(2)
        strategic_score = (x_norm + y_norm) / 2 - dist_to_axis * 0.5
        strategic_scores.append(strategic_score)
    
    strategic_scores = np.array(strategic_scores)
    
    colors_est = []
    for score in strategic_scores:
        if score > np.percentile(strategic_scores, 75):
            colors_est.append('#CC0000')
        elif score > np.percentile(strategic_scores, 50):
            colors_est.append('#FF6600')
        elif score > np.percentile(strategic_scores, 25):
            colors_est.append('#3388BB')
        else:
            colors_est.append('#888888')
    
    sizes_est = 50 + (strategic_scores - strategic_scores.min()) / (strategic_scores.max() - strategic_scores.min()) * 150
    
    ax_est.scatter(dep_tot, mot_tot, c=colors_est, s=sizes_est, alpha=0.7, edgecolors='black')
    ax_est.plot([0, max_dep_norm], [0, max_mot_norm], 'r--', linewidth=3, alpha=0.8, label='Eje')
    
    strategic_indices = np.argsort(strategic_scores)[-min(15, len(nombres)):]
    for idx in strategic_indices:
        ax_est.annotate(
            nombres[idx][:25],
            (dep_tot[idx], mot_tot[idx]),
            xytext=(8, 8), textcoords='offset points',
            fontsize=9, fontweight='bold',
            bbox=dict(boxstyle="round", facecolor="yellow", alpha=0.85),
            arrowprops=dict(arrowstyle='->', color='orange')
        )
    
    ax_est.set_xlabel("Dependencia", fontweight='bold', fontsize=14)
    ax_est.set_ylabel("Motricidad", fontweight='bold', fontsize=14)
    ax_est.set_title(f"EJE ESTRATÉGICO (α={alpha}, K={K_max})", fontweight='bold', fontsize=16)
    ax_est.grid(True, alpha=0.3)
    ax_est.legend(fontsize=12)
    
    st.pyplot(fig_est)

# TAB 4
with tab4:
    st.markdown("### 🔬 Análisis de Estabilidad")
    
    col1, col2 = st.columns(2)
    with col1:
        alphas_test = st.multiselect("α:", [0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8], default=[0.3, 0.5, 0.7])
    with col2:
        Ks_test = st.multiselect("K:", list(range(2, 11)), default=[3, 6, 9])
    
    if st.button("🔄 Ejecutar"):
        with st.spinner("Calculando..."):
            df_stability = analyze_stability(M, alphas_test, Ks_test)
            
            for i in range(1, 6):
                df_stability[f'Var_Top{i}'] = df_stability[f'top_{i}'].apply(lambda idx: nombres[idx])
            
            st.success(f"✅ {len(df_stability)} configuraciones probadas")
            
            display_cols = ['alpha', 'K'] + [f'Var_Top{i}' for i in range(1, 6)]
            st.dataframe(df_stability[display_cols], use_container_width=True)

# TAB 5
with tab5:
    st.markdown("### 📊 Top 15")
    
    fig_bar, ax_bar = plt.subplots(figsize=(14, 8))
    
    top_15_idx = order[:15]
    top_15_vars = [nombres[i] for i in top_15_idx]
    top_15_mot = mot_tot[top_15_idx]
    
    colors_bar = []
    for var in top_15_vars:
        clf = df_all.loc[var, 'Clasificación']
        if clf == 'Crítico/inestable':
            colors_bar.append('#1166CC')
        elif clf == 'Determinantes':
            colors_bar.append('#FF4444')
        elif clf == 'Variables resultado':
            colors_bar.append('#66BBFF')
        else:
            colors_bar.append('#FF9944')
    
    y_pos = np.arange(len(top_15_vars))
    ax_bar.barh(y_pos, top_15_mot, color=colors_bar, edgecolor='black')
    ax_bar.set_yticks(y_pos)
    ax_bar.set_yticklabels(top_15_vars)
    ax_bar.invert_yaxis()
    ax_bar.set_xlabel("Motricidad", fontweight='bold')
    ax_bar.set_title(f"Top 15 (α={alpha}, K={K_max})", fontweight='bold')
    ax_bar.grid(axis='x', alpha=0.3)
    
    for i, val in enumerate(top_15_mot):
        ax_bar.text(val, i, f' {val:.0f}', va='center', fontsize=9, fontweight='bold')
    
    st.pyplot(fig_bar)
# ============================================================
# GRÁFICO DE INFLUENCIAS INDIRECTAS (GRAFO DE RED)
# ============================================================

st.markdown("---")
st.markdown("#### 🕸️ Gráfico de Influencias Indirectas")
st.caption("Visualización de red: propagación de influencias entre variables")

# Instalar networkx si es necesario
try:
    import networkx as nx
except ImportError:
    st.error("⚠️ Esta visualización requiere networkx. Instala con: pip install networkx")
    st.stop()

# Calcular matriz de influencias indirectas (ya está calculada)
M_ind_matrix = M_tot - M

# Parámetros de visualización
col_g1, col_g2, col_g3 = st.columns(3)

with col_g1:
    umbral_minimo = st.slider(
        "Umbral mínimo de influencia",
        min_value=0.0,
        max_value=float(np.percentile(M_ind_matrix[M_ind_matrix > 0], 50)),
        value=float(np.percentile(M_ind_matrix[M_ind_matrix > 0], 25)),
        step=0.1,
        help="Mostrar solo influencias indirectas mayores a este valor"
    )

with col_g2:
    max_nodos = st.slider(
        "Máximo de nodos a mostrar",
        min_value=10,
        max_value=min(50, len(nombres)),
        value=min(25, len(nombres)),
        help="Limitar cantidad de nodos para mejor visualización"
    )

with col_g3:
    layout_tipo = st.selectbox(
        "Tipo de layout",
        options=['spring', 'circular', 'kamada_kawai'],
        index=0,
        help="Algoritmo de posicionamiento de nodos"
    )

# Seleccionar top N nodos por motricidad indirecta
top_nodos_idx = np.argsort(mot_ind)[-max_nodos:]
nombres_seleccionados = [nombres[i] for i in top_nodos_idx]

# Crear grafo dirigido
G = nx.DiGraph()

# Agregar nodos
for i, var in enumerate(nombres_seleccionados):
    idx_original = nombres.index(var)
    G.add_node(var, 
               motricidad=mot_ind[idx_original],
               dependencia=dep_ind[idx_original],
               clasificacion=df_all.loc[var, 'Clasificación'])

# Clasificar influencias indirectas en categorías
influencias_todas = M_ind_matrix[M_ind_matrix > umbral_minimo]

if len(influencias_todas) > 0:
    # Percentiles para clasificación
    p25 = np.percentile(influencias_todas, 25)
    p50 = np.percentile(influencias_todas, 50)
    p75 = np.percentile(influencias_todas, 75)
    p90 = np.percentile(influencias_todas, 90)
    
    # Agregar aristas con clasificación
    for i, var_origen in enumerate(nombres_seleccionados):
        idx_i = nombres.index(var_origen)
        for j, var_destino in enumerate(nombres_seleccionados):
            if i != j:  # No auto-loops
                idx_j = nombres.index(var_destino)
                influencia = M_ind_matrix[idx_i, idx_j]
                
                if influencia > umbral_minimo:
                    # Clasificar influencia
                    if influencia > p90:
                        categoria = 'muy_importante'
                        color = '#CC0000'  # Rojo oscuro
                        ancho = 3.0
                    elif influencia > p75:
                        categoria = 'importante'
                        color = '#FF6600'  # Naranja
                        ancho = 2.5
                    elif influencia > p50:
                        categoria = 'media'
                        color = '#FFAA00'  # Amarillo-naranja
                        ancho = 2.0
                    elif influencia > p25:
                        categoria = 'debil'
                        color = '#88BBFF'  # Azul claro
                        ancho = 1.5
                    else:
                        categoria = 'muy_debil'
                        color = '#CCCCCC'  # Gris
                        ancho = 1.0
                    
                    G.add_edge(var_origen, var_destino,
                             weight=influencia,
                             categoria=categoria,
                             color=color,
                             ancho=ancho)

# Verificar que hay aristas
if G.number_of_edges() == 0:
    st.warning(f"⚠️ No hay influencias indirectas mayores al umbral {umbral_minimo:.2f}. Reduce el umbral.")
else:
    # Crear figura
    fig_grafo, ax_grafo = plt.subplots(figsize=(20, 16))



      # Layout del grafo con mejor espaciado
    espaciado_factor = 8.0  # Ajusta este valor (2.0 = menos espacio, 8.0 = mucho espacio)

    if layout_tipo == 'spring':
        pos = nx.spring_layout(
            G, 
            k=espaciado_factor,    # Distancia ideal entre nodos
            iterations=150,        # Más iteraciones para mejor distribución  
            seed=42
        )
    elif layout_tipo == 'circular':
        pos = nx.circular_layout(G, scale=espaciado_factor)
    else:  # kamada_kawai
        pos = nx.kamada_kawai_layout(G, scale=espaciado_factor)

    # Expandir posiciones adicional si es necesario
    expansion_extra = 1.2
    pos = {node: (x * expansion_extra, y * expansion_extra) for node, (x, y) in pos.items()}

   



    
    # Colores de nodos según clasificación
    node_colors = []
    node_sizes = []
    for node in G.nodes():
        clasificacion = G.nodes[node]['clasificacion']
        motricidad = G.nodes[node]['motricidad']
        
        if clasificacion == 'Crítico/inestable':
            node_colors.append('#1166CC')
        elif clasificacion == 'Determinantes':
            node_colors.append('#FF4444')
        elif clasificacion == 'Variables resultado':
            node_colors.append('#66BBFF')
        else:
            node_colors.append('#FF9944')
        
        # Tamaño proporcional a motricidad indirecta
        node_sizes.append(100 + motricidad * 3)
    
    # Dibujar aristas por categoría
    categorias_aristas = {
        'muy_importante': {'edges': [], 'color': '#CC0000', 'ancho': 3.0, 'alpha': 0.8, 'label': 'Muy importantes'},
        'importante': {'edges': [], 'color': '#FF6600', 'ancho': 2.5, 'alpha': 0.7, 'label': 'Importantes'},
        'media': {'edges': [], 'color': '#FFAA00', 'ancho': 2.0, 'alpha': 0.6, 'label': 'Medias'},
        'debil': {'edges': [], 'color': '#88BBFF', 'ancho': 1.5, 'alpha': 0.5, 'label': 'Débiles'},
        'muy_debil': {'edges': [], 'color': '#CCCCCC', 'ancho': 1.0, 'alpha': 0.4, 'label': 'Muy débiles'}
    }
    
    # Agrupar aristas por categoría
    for (u, v, data) in G.edges(data=True):
        categoria = data['categoria']
        categorias_aristas[categoria]['edges'].append((u, v))
    
    # Dibujar aristas por categoría (de menos a más importante para superposición correcta)
    for cat in ['muy_debil', 'debil', 'media', 'importante', 'muy_importante']:
        if categorias_aristas[cat]['edges']:
            nx.draw_networkx_edges(
                G, pos,
                edgelist=categorias_aristas[cat]['edges'],
                edge_color=categorias_aristas[cat]['color'],
                width=categorias_aristas[cat]['ancho'],
                alpha=categorias_aristas[cat]['alpha'],
                arrows=True,
                arrowsize=15,
                arrowstyle='->',
                connectionstyle='arc3,rad=0.1',
                ax=ax_grafo
            )
    
    # Dibujar nodos
    nx.draw_networkx_nodes(
        G, pos,
        node_color=node_colors,
        node_size=node_sizes,
        alpha=0.9,
        edgecolors='black',
        linewidths=2,
        ax=ax_grafo
    )
    
    # Etiquetas de nodos
    labels = {node: node[:20] for node in G.nodes()}  # Truncar nombres largos
    nx.draw_networkx_labels(
        G, pos,
        labels=labels,
        font_size=8,
        font_weight='bold',
        font_color='black',
        ax=ax_grafo
    )
    
    # Título y leyenda
    ax_grafo.set_title(
        f"GRÁFICO DE INFLUENCIAS INDIRECTAS (α={alpha}, K={K_max})\n"
        f"Nodos: {G.number_of_nodes()} | Conexiones: {G.number_of_edges()}",
        fontweight='bold',
        fontsize=16,
        pad=20
    )
    
    # Leyenda personalizada
    from matplotlib.lines import Line2D
    legend_elements = []
    
    # Leyenda de aristas (influencias)
    for cat in ['muy_importante', 'importante', 'media', 'debil', 'muy_debil']:
        if categorias_aristas[cat]['edges']:
            legend_elements.append(
                Line2D([0], [0], color=categorias_aristas[cat]['color'],
                      linewidth=categorias_aristas[cat]['ancho'],
                      label=f"{categorias_aristas[cat]['label']} ({len(categorias_aristas[cat]['edges'])})")
            )
    
    legend_elements.append(Line2D([0], [0], color='white', linewidth=0, label=''))  # Separador
    
    # Leyenda de nodos (clasificación)
    legend_elements.extend([
        Line2D([0], [0], marker='o', color='w', markerfacecolor='#FF4444', 
               markersize=10, label='Determinantes'),
        Line2D([0], [0], marker='o', color='w', markerfacecolor='#1166CC', 
               markersize=10, label='Crítico/inestable'),
        Line2D([0], [0], marker='o', color='w', markerfacecolor='#66BBFF', 
               markersize=10, label='Variables resultado'),
        Line2D([0], [0], marker='o', color='w', markerfacecolor='#FF9944', 
               markersize=10, label='Autónomas')
    ])
    
    ax_grafo.legend(handles=legend_elements, loc='upper left', fontsize=10, 
                   frameon=True, shadow=True, title='Leyenda')
    
    ax_grafo.axis('off')
    ax_grafo.margins(0.1)
    
    st.pyplot(fig_grafo)
    
    # Estadísticas del grafo
    st.markdown("#### 📊 Estadísticas del Grafo de Influencias Indirectas")
    
    col_e1, col_e2, col_e3, col_e4 = st.columns(4)
    
    with col_e1:
        st.metric("Nodos (Variables)", G.number_of_nodes())
    
    with col_e2:
        st.metric("Conexiones (Influencias)", G.number_of_edges())
    
    with col_e3:
        densidad = nx.density(G)
        st.metric("Densidad de Red", f"{densidad:.3f}")
    
    with col_e4:
        grado_medio = sum(dict(G.degree()).values()) / G.number_of_nodes()
        st.metric("Grado Medio", f"{grado_medio:.1f}")
    
    # Top 10 variables más conectadas (mayor grado de salida)
    st.markdown("#### 🎯 Top 10 Variables con Mayor Influencia Indirecta (Grado de Salida)")
    
    out_degree = dict(G.out_degree())
    top_influencers = sorted(out_degree.items(), key=lambda x: x[1], reverse=True)[:10]
    
    df_influencers = pd.DataFrame({
        'Variable': [var for var, deg in top_influencers],
        'Conexiones_Salida': [deg for var, deg in top_influencers],
        'Motricidad_Indirecta': [mot_ind[nombres.index(var)] for var, deg in top_influencers],
        'Clasificación': [df_all.loc[var, 'Clasificación'] for var, deg in top_influencers]
    })
    
    st.dataframe(
        df_influencers.style.background_gradient(subset=['Conexiones_Salida'], cmap='Reds'),
        use_container_width=True
    )
    
    # Botón de descarga
    img_grafo = io.BytesIO()
    fig_grafo.savefig(img_grafo, format='png', dpi=300, bbox_inches='tight', facecolor='white')
    img_grafo.seek(0)
    st.download_button(
        label="📥 Descargar Gráfico de Influencias Indirectas (PNG)",
        data=img_grafo,
        file_name=f"micmac_influencias_indirectas_a{alpha}_k{K_max}.png",
        mime="image/png"
    )
    
    # Interpretación
    with st.expander("ℹ️ Interpretación del Gráfico de Influencias Indirectas"):
        st.markdown("""
        ### ¿Cómo interpretar este gráfico?
        
        **Nodos (Variables):**
        - **Tamaño:** Proporcional a la motricidad indirecta de la variable
        - **Color:** Según clasificación MICMAC (Determinantes, Críticas, Resultado, Autónomas)
        
        **Aristas (Influencias Indirectas):**
        - **Color y grosor:** Indican la intensidad de la influencia indirecta
        - **Muy importantes (rojo oscuro):** Influencias indirectas muy fuertes (>percentil 90)
        - **Importantes (naranja):** Influencias indirectas fuertes (percentil 75-90)
        - **Medias (amarillo-naranja):** Influencias indirectas moderadas (percentil 50-75)
        - **Débiles (azul claro):** Influencias indirectas bajas (percentil 25-50)
        - **Muy débiles (gris):** Influencias indirectas mínimas (< percentil 25)
        
        **Análisis de Red:**
        - **Densidad:** Indica qué tan interconectado está el sistema (0 = sin conexiones, 1 = totalmente conectado)
        - **Grado medio:** Número promedio de conexiones por variable
        - **Hubs (concentradores):** Variables con muchas conexiones de salida → alto poder de influencia indirecta
        - **Puentes:** Variables que conectan grupos → facilitadores de propagación de cambios
        
        ### Interpretación Estratégica
        
        Este gráfico revela las **cadenas de influencia** que no son evidentes en el análisis directo:
        - Variables que influyen **indirectamente** a través de intermediarios
        - Efectos **multiplicadores** y **cascada** en el sistema
        - Variables que funcionan como **transmisores** o **amplificadores**
        
        **Recomendación:** Presta atención especial a las variables con:
        - Muchas conexiones de salida (alto grado) → puntos de intervención para cambios sistémicos
        - Conexiones rojas/naranjas → canales de influencia muy fuertes
        - Posición central en el grafo → facilitadores clave de la dinámica del sistema
        """)

# TAB 6
with tab6:

    # ============================================================
    # GENERADOR DE INFORME DE INTELIGENCIA CON PDF COMPLETO
    # ============================================================

    st.markdown("---")
    st.markdown("### 📄 Generador de Informe de Inteligencia")
    st.caption("Crea un informe PDF profesional con todos los gráficos y análisis")

    if st.button("🎯 Generar Informe Completo (PDF)", type="primary"):
        with st.spinner("Generando informe PDF con todos los gráficos... Por favor espera."):
            
            # Importar PdfPages para PDF multipágina
            from matplotlib.backends.backend_pdf import PdfPages
            import tempfile
            
            # Análisis automático de resultados
            top_5_motoras = ranking_vars[:5]
            
            # RECREAR VARIABLES NECESARIAS PARA EL INFORME
            # Recrear clasificación por cuadrantes
            ref_x = np.median(mot_tot)
            ref_y = np.median(dep_tot)
            
            labels_cuadrante = []
            colors = []
            color_map = {
                "Determinantes":      "#FF4444",
                "Crítico/inestable":  "#1166CC",
                "Variables resultado":"#66BBFF",
                "Autónomas":          "#FF9944",
            }
            
            for xi, yi in zip(mot_tot, dep_tot):
                if   xi >= ref_x and yi <  ref_y: 
                    labels_cuadrante.append("Determinantes")
                    colors.append(color_map["Determinantes"])
                elif xi >= ref_x and yi >= ref_y: 
                    labels_cuadrante.append("Crítico/inestable")
                    colors.append(color_map["Crítico/inestable"])
                elif xi <  ref_x and yi >= ref_y: 
                    labels_cuadrante.append("Variables resultado")
                    colors.append(color_map["Variables resultado"])
                else:                              
                    labels_cuadrante.append("Autónomas")
                    colors.append(color_map["Autónomas"])
            
            # Recrear scores estratégicos
            x_norm = mot_tot / (mot_tot.max() if mot_tot.max()!=0 else 1.0)
            y_norm = dep_tot / (dep_tot.max() if dep_tot.max()!=0 else 1.0)
            dist = np.abs(y_norm - x_norm) / np.sqrt(2)
            strategic_scores = (x_norm + y_norm)/2 - dist
            
            # Variables estratégicas top 3
            top_3_estrategicas = [nombres[i] for i in np.argsort(strategic_scores)[-3:]][::-1]
            
            # Contar variables por cuadrante
            count_determinantes = sum(1 for label in labels_cuadrante if label == 'Determinantes')
            count_criticas = sum(1 for label in labels_cuadrante if label == 'Crítico/inestable')
            count_resultado = sum(1 for label in labels_cuadrante if label == 'Variables resultado')
            count_autonomas = sum(1 for label in labels_cuadrante if label == 'Autónomas')
            
            # Variables críticas por motricidad
            vars_alta_motricidad = [nombres[i] for i in range(len(nombres)) if mot_tot[i] > np.percentile(mot_tot, 90)]
            vars_alta_dependencia = [nombres[i] for i in range(len(nombres)) if dep_tot[i] > np.percentile(dep_tot, 90)]
            
            # Colores para eje estratégico
            p80, p60, p40 = np.percentile(strategic_scores, [80, 60, 40])
            col_est = []
            for s in strategic_scores:
                if s > p80:   col_est.append("#CC0000")
                elif s > p60: col_est.append("#FF6600")
                elif s > p40: col_est.append("#3388BB")
                else:         col_est.append("#888888")
            
            sizes_est = 50 + 100*(strategic_scores - strategic_scores.min())/(
                (strategic_scores.max() - strategic_scores.min()) if strategic_scores.max()>strategic_scores.min() else 1.0
            )
            
            # Crear archivo temporal para PDF
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                with PdfPages(tmp_file.name) as pdf:
                    
                    # ========================================
                    # PÁGINA 1: PORTADA DEL INFORME
                    # ========================================
                    fig_portada = plt.figure(figsize=(8.5, 11))
                    ax_portada = fig_portada.add_subplot(111)
                    ax_portada.axis('off')
                    
                    # Título principal
                    ax_portada.text(0.5, 0.85, 'INFORME DE INTELIGENCIA ESTRATÉGICA', 
                                   fontsize=24, fontweight='bold', ha='center', va='center',
                                   bbox=dict(boxstyle="round,pad=1", facecolor="lightblue", alpha=0.8))
                    
                    # Subtítulo
                    ax_portada.text(0.5, 0.78, 'Análisis Estructural MICMAC', 
                                   fontsize=18, ha='center', va='center')
                    
                    # Información del análisis
                    fecha_actual = datetime.now().strftime("%d de %B de %Y")
                    info_text = f"""Sistema Analizado: {len(nombres)} Variables
Parámetros: α={alpha}, K={K_max}
Fecha de Análisis: {fecha_actual}

Generado por: Sistema MICMAC Interactivo v3.5
Desarrollado por: Martín Pratto"""
                    
                    ax_portada.text(0.5, 0.5, info_text, fontsize=12, ha='center', va='center',
                                   bbox=dict(boxstyle="round,pad=0.5", facecolor="lightyellow", alpha=0.7))
                    
                    # Resumen ejecutivo
                    resumen = f"""HALLAZGOS PRINCIPALES:

• {count_criticas} Variables Críticas/Inestables ({count_criticas/len(nombres)*100:.1f}%)
• {count_determinantes} Variables Determinantes ({count_determinantes/len(nombres)*100:.1f}%)
• {count_resultado} Variables Resultado ({count_resultado/len(nombres)*100:.1f}%)

Variables Más Estratégicas:
1. {top_3_estrategicas[0]}
2. {top_3_estrategicas[1]}
3. {top_3_estrategicas[2]}"""
                    
                    ax_portada.text(0.5, 0.2, resumen, fontsize=11, ha='center', va='center',
                                   bbox=dict(boxstyle="round,pad=0.5", facecolor="lightgreen", alpha=0.7))
                    
                    pdf.savefig(fig_portada, bbox_inches='tight')
                    plt.close(fig_portada)
                    
                    # ========================================
                    # PÁGINA 2: MAPA MICMAC (TOTAL)
                    # ========================================
                    fig_mapa_pdf = plt.figure(figsize=(12, 9))
                    ax_mapa_pdf = fig_mapa_pdf.add_subplot(111)
                    
                    # Recrear mapa MICMAC
                    sc = ax_mapa_pdf.scatter(mot_tot, dep_tot, c=colors, s=120, alpha=0.85, edgecolors='black', linewidth=1.0)
                    ax_mapa_pdf.axvline(ref_x, color='black', linestyle='--', linewidth=1.2, alpha=0.8)
                    ax_mapa_pdf.axhline(ref_y, color='black', linestyle='--', linewidth=1.2, alpha=0.8)
                    
                    # Etiquetas (solo top variables)
                    top_indices_pdf = np.argsort(-mot_tot)[:15]  # Top 15
                    for i in top_indices_pdf:
                        ax_mapa_pdf.text(mot_tot[i], dep_tot[i], f" {nombres[i][:15]}", fontsize=8)
                    
                    ax_mapa_pdf.set_xlabel("Motricidad (Total)", fontweight='bold', fontsize=12)
                    ax_mapa_pdf.set_ylabel("Dependencia (Total)", fontweight='bold', fontsize=12)
                    ax_mapa_pdf.set_title(f"MAPA MICMAC TOTAL — α={alpha}, K={K_max}", fontweight='bold', fontsize=14)
                    
                    # Leyenda
                    handles_pdf = [
                        plt.Line2D([0],[0], marker='o', color='w', markerfacecolor=color_map["Determinantes"], markersize=10, label='Determinantes'),
                        plt.Line2D([0],[0], marker='o', color='w', markerfacecolor=color_map["Crítico/inestable"], markersize=10, label='Crítico/inestable'),
                        plt.Line2D([0],[0], marker='o', color='w', markerfacecolor=color_map["Variables resultado"], markersize=10, label='Variables resultado'),
                        plt.Line2D([0],[0], marker='o', color='w', markerfacecolor=color_map["Autónomas"], markersize=10, label='Autónomas'),
                    ]
                    ax_mapa_pdf.legend(handles=handles_pdf, loc='upper left', frameon=True)
                    ax_mapa_pdf.grid(True, alpha=0.25)
                    
                    pdf.savefig(fig_mapa_pdf, bbox_inches='tight')
                    plt.close(fig_mapa_pdf)
                    
                    # ========================================
                    # PÁGINA 3: EJE DE ESTRATEGIA
                    # ========================================
                    fig_eje_pdf = plt.figure(figsize=(11, 9))
                    ax_eje_pdf = fig_eje_pdf.add_subplot(111)
                    
                    # Recrear eje de estrategia
                    ax_eje_pdf.scatter(mot_tot, dep_tot, c=col_est, s=sizes_est, alpha=0.85, edgecolors='black', linewidth=1.0)
                    ax_eje_pdf.plot([0, mot_tot.max()], [0, dep_tot.max()], 'r--', lw=2, label='Eje de estrategia')
                    
                    # Etiquetas top estratégicas
                    idx_top_est_pdf = np.argsort(strategic_scores)[-10:]  # Top 10 estratégicas
                    for i in idx_top_est_pdf:
                        ax_eje_pdf.text(mot_tot[i], dep_tot[i], f" {nombres[i][:15]}", fontsize=8)
                    
                    ax_eje_pdf.set_xlabel("Motricidad (Total)", fontweight='bold', fontsize=12)
                    ax_eje_pdf.set_ylabel("Dependencia (Total)", fontweight='bold', fontsize=12)
                    ax_eje_pdf.set_title("EJE DE ESTRATEGIA", fontweight='bold', fontsize=14)
                    ax_eje_pdf.legend(loc='upper left')
                    ax_eje_pdf.grid(True, alpha=0.25)
                    
                    pdf.savefig(fig_eje_pdf, bbox_inches='tight')
                    plt.close(fig_eje_pdf)
                    
                    # Resto del código igual (página 4, 5, 6, 7)...
                    # [CONTINÚA CON EL RESTO DEL CÓDIGO ANTERIOR]

                    
                    # ========================================
                    # PÁGINA 2: MAPA MICMAC (TOTAL)
                    # ========================================
                    fig_mapa_pdf = plt.figure(figsize=(12, 9))
                    ax_mapa_pdf = fig_mapa_pdf.add_subplot(111)
                    
                    # Recrear mapa MICMAC
                    sc = ax_mapa_pdf.scatter(mot_tot, dep_tot, c=colors, s=120, alpha=0.85, edgecolors='black', linewidth=1.0)
                    ax_mapa_pdf.axvline(ref_x, color='black', linestyle='--', linewidth=1.2, alpha=0.8)
                    ax_mapa_pdf.axhline(ref_y, color='black', linestyle='--', linewidth=1.2, alpha=0.8)

                    # Etiquetas (solo top variables)
                    top_indices_pdf = np.argsort(-mot_tot)[:15]  # Top 15
                    for i in top_indices_pdf:
                      ax_mapa_pdf.text(mot_tot[i], dep_tot[i], f" {nombres[i][:15]}", fontsize=8)
                    
                                        
                    ax_mapa_pdf.set_xlabel("Motricidad (Total)", fontweight='bold', fontsize=12)
                    ax_mapa_pdf.set_ylabel("Dependencia (Total)", fontweight='bold', fontsize=12)
                    ax_mapa_pdf.set_title(f"MAPA MICMAC TOTAL — α={alpha}, K={K_max}", fontweight='bold', fontsize=14)
                    
                    # Leyenda
                    handles_pdf = [
                        plt.Line2D([0],[0], marker='o', color='w', markerfacecolor=color_map["Determinantes"], markersize=10, label='Determinantes'),
                        plt.Line2D([0],[0], marker='o', color='w', markerfacecolor=color_map["Crítico/inestable"], markersize=10, label='Crítico/inestable'),
                        plt.Line2D([0],[0], marker='o', color='w', markerfacecolor=color_map["Variables resultado"], markersize=10, label='Variables resultado'),
                        plt.Line2D([0],[0], marker='o', color='w', markerfacecolor=color_map["Autónomas"], markersize=10, label='Autónomas'),
                    ]
                    ax_mapa_pdf.legend(handles=handles_pdf, loc='upper left', frameon=True)
                    ax_mapa_pdf.grid(True, alpha=0.25)
                    
                    pdf.savefig(fig_mapa_pdf, bbox_inches='tight')
                    plt.close(fig_mapa_pdf)
                    
                    # ========================================
                    # PÁGINA 3: EJE DE ESTRATEGIA
                    # ========================================
                    fig_eje_pdf = plt.figure(figsize=(11, 9))
                    ax_eje_pdf = fig_eje_pdf.add_subplot(111)
                    
                    # Recrear eje de estrategia
                    ax_eje_pdf.scatter(X, Y, c=col_est, s=sizes_est, alpha=0.85, edgecolors='black', linewidth=1.0)
                    ax_eje_pdf.plot([0, X.max()], [0, Y.max()], 'r--', lw=2, label='Eje de estrategia')
                    
                    # Etiquetas top estratégicas
                    idx_top_est_pdf = np.argsort(strategic_scores)[-10:]  # Top 10 estratégicas
                    for i in idx_top_est_pdf:
                        ax_eje_pdf.text(X[i], Y[i], f" {nombres[i][:15]}", fontsize=8)
                    
                    ax_eje_pdf.set_xlabel("Motricidad (Total)", fontweight='bold', fontsize=12)
                    ax_eje_pdf.set_ylabel("Dependencia (Total)", fontweight='bold', fontsize=12)
                    ax_eje_pdf.set_title("EJE DE ESTRATEGIA", fontweight='bold', fontsize=14)
                    ax_eje_pdf.legend(loc='upper left')
                    ax_eje_pdf.grid(True, alpha=0.25)
                    
                    pdf.savefig(fig_eje_pdf, bbox_inches='tight')
                    plt.close(fig_eje_pdf)
                    
                    # ========================================
                    # PÁGINA 4: RANKING DE VARIABLES (BARRAS)
                    # ========================================
                    fig_barras_pdf = plt.figure(figsize=(14, 10))
                    ax_barras_pdf = fig_barras_pdf.add_subplot(111)
                    
                    # Top 20 variables
                    df_top20 = df_rank.head(20)
                    bars = ax_barras_pdf.barh(range(len(df_top20)), df_top20["Motricidad_total"], color='steelblue', alpha=0.8)
                    ax_barras_pdf.set_yticks(range(len(df_top20)))
                    ax_barras_pdf.set_yticklabels(df_top20["Variable"], fontsize=10)
                    ax_barras_pdf.set_xlabel("Motricidad Total", fontweight='bold', fontsize=12)
                    ax_barras_pdf.set_title("TOP 20 VARIABLES POR MOTRICIDAD TOTAL", fontweight='bold', fontsize=14)
                    ax_barras_pdf.grid(True, alpha=0.25, axis='x')
                    
                    # Invertir orden para que el top esté arriba
                    ax_barras_pdf.invert_yaxis()
                    
                    pdf.savefig(fig_barras_pdf, bbox_inches='tight')
                    plt.close(fig_barras_pdf)
                    
                    # ========================================
                    # PÁGINA 5: GRAFO DE RED (si existe)
                    # ========================================
                    try:
                        if 'G' in locals() and G.number_of_edges() > 0:
                            fig_grafo_pdf = plt.figure(figsize=(16, 12))
                            ax_grafo_pdf = fig_grafo_pdf.add_subplot(111)
                            
                            # Recrear layout del grafo
                            pos_pdf = nx.spring_layout(G, k=5.0, iterations=100, seed=42)
                            expansion_pdf = 1.2
                            pos_pdf = {node: (x * expansion_pdf, y * expansion_pdf) for node, (x, y) in pos_pdf.items()}
                            
                            # Colores de nodos
                            node_colors_pdf = []
                            node_sizes_pdf = []
                            for node in G.nodes():
                                clasificacion = G.nodes[node]['clasificacion']
                                motricidad_node = G.nodes[node]['motricidad']
                                
                                if clasificacion == 'Crítico/inestable':
                                    node_colors_pdf.append('#1166CC')
                                elif clasificacion == 'Determinantes':
                                    node_colors_pdf.append('#FF4444')
                                elif clasificacion == 'Variables resultado':
                                    node_colors_pdf.append('#66BBFF')
                                else:
                                    node_colors_pdf.append('#FF9944')
                                
                                node_sizes_pdf.append(100 + motricidad_node * 2)
                            
                            # Dibujar aristas por categoría
                            for cat in ['muy_debil', 'debil', 'media', 'importante', 'muy_importante']:
                                if cat in categorias_aristas and categorias_aristas[cat]['edges']:
                                    nx.draw_networkx_edges(
                                        G, pos_pdf,
                                        edgelist=categorias_aristas[cat]['edges'],
                                        edge_color=categorias_aristas[cat]['color'],
                                        width=categorias_aristas[cat]['ancho'],
                                        alpha=categorias_aristas[cat]['alpha'],
                                        arrows=True,
                                        arrowsize=10,
                                        ax=ax_grafo_pdf
                                    )
                            
                            # Dibujar nodos
                            nx.draw_networkx_nodes(
                                G, pos_pdf,
                                node_color=node_colors_pdf,
                                node_size=node_sizes_pdf,
                                alpha=0.8,
                                edgecolors='black',
                                linewidths=1,
                                ax=ax_grafo_pdf
                            )
                            
                            # Etiquetas
                            labels_pdf = {node: node[:12] for node in G.nodes()}
                            nx.draw_networkx_labels(G, pos_pdf, labels_pdf, font_size=7, ax=ax_grafo_pdf)
                            
                            ax_grafo_pdf.set_title(f"GRAFO DE INFLUENCIAS INDIRECTAS\nNodos: {G.number_of_nodes()} | Conexiones: {G.number_of_edges()}", 
                                                  fontweight='bold', fontsize=14)
                            ax_grafo_pdf.axis('off')
                            
                            pdf.savefig(fig_grafo_pdf, bbox_inches='tight')
                            plt.close(fig_grafo_pdf)
                    except:
                        pass  # Si no hay grafo, saltear esta página
                    
                    # ========================================
                    # PÁGINA 6: ANÁLISIS TEXTUAL
                    # ========================================
                    fig_texto = plt.figure(figsize=(8.5, 11))
                    ax_texto = fig_texto.add_subplot(111)
                    ax_texto.axis('off')
                    
                    informe_texto = f"""
ANÁLISIS DE VARIABLES MOTORAS

Top 5 Variables con Mayor Influencia Sistémica:
1. {top_5_motoras[0]} - Motricidad: {mot_tot[order[0]]:.0f}
2. {top_5_motoras[1]} - Motricidad: {mot_tot[order[1]]:.0f}
3. {top_5_motoras[2]} - Motricidad: {mot_tot[order[2]]:.0f}
4. {top_5_motoras[3]} - Motricidad: {mot_tot[order[3]]:.0f}
5. {top_5_motoras[4]} - Motricidad: {mot_tot[order[4]]:.0f}

CLASIFICACIÓN SISTÉMICA

• Variables Críticas/Inestables: {count_criticas} ({count_criticas/len(nombres)*100:.1f}%)
• Variables Determinantes: {count_determinantes} ({count_determinantes/len(nombres)*100:.1f}%)
• Variables Resultado: {count_resultado} ({count_resultado/len(nombres)*100:.1f}%)
• Variables Autónomas: {count_autonomas} ({count_autonomas/len(nombres)*100:.1f}%)

RECOMENDACIONES ESTRATÉGICAS

PRIORIDAD ALTA - Acción Inmediata:
1. Focalizar recursos en las {count_determinantes} variables determinantes
2. Gestión de variables críticas que pueden generar efectos impredecibles

PRIORIDAD MEDIA - Planificación Táctica:
3. Monitoreo de variables resultado como sistema de alerta
4. Optimización del eje estratégico

INDICADORES CLAVE:
• Motricidad Concentrada: {(mot_tot[order[0]]/mot_tot.sum()*100):.2f}%
• Ratio Variables Críticas: {count_criticas/len(nombres):.3f}
• Dependencia Media: {dep_tot.mean():.2f}

METODOLOGÍA APLICADA:
• Algoritmo: MICMAC extendido con parámetros α={alpha}, K={K_max}
• Variables analizadas: {len(nombres)}
• Fecha de análisis: {fecha_actual}
                    """
                    
                    ax_texto.text(0.05, 0.95, informe_texto, fontsize=10, ha='left', va='top',
                                 transform=ax_texto.transAxes, fontfamily='monospace')
                    
                    pdf.savefig(fig_texto, bbox_inches='tight')
                    plt.close(fig_texto)
                    
                    # ========================================
                    # PÁGINA 7: TABLAS DE DATOS
                    # ========================================
                    fig_tabla = plt.figure(figsize=(11, 8.5))
                    ax_tabla = fig_tabla.add_subplot(111)
                    ax_tabla.axis('off')
                    
                    # Crear tabla con top 15 variables
                    tabla_data = []
                    for i in range(min(15, len(order))):  # Evitar error si hay menos de 15 variables
                        idx = order[i]
                        tabla_data.append([
                            i + 1,
                            nombres[idx][:25],
                            f"{mot_tot[idx]:.0f}",
                            f"{dep_tot[idx]:.0f}",
                            labels_cuadrante[idx]
                        ])
                    
                    # Crear tabla
                    tabla = ax_tabla.table(
                        cellText=tabla_data,
                        colLabels=['Pos', 'Variable', 'Motricidad', 'Dependencia', 'Clasificación'],
                        cellLoc='left',
                        loc='center',
                        colWidths=[0.08, 0.45, 0.15, 0.15, 0.17]
                    )
                    tabla.auto_set_font_size(False)
                    tabla.set_fontsize(9)
                    tabla.scale(1, 1.5)
                    
                    # Colorear encabezados
                    for i in range(5):
                        tabla[(0, i)].set_facecolor('#4CAF50')
                        tabla[(0, i)].set_text_props(weight='bold', color='white')
                    
                    # Colorear filas alternadas
                    for i in range(1, len(tabla_data) + 1):
                        for j in range(5):
                            if i % 2 == 0:
                                tabla[(i, j)].set_facecolor('#f0f0f0')
                    
                    ax_tabla.set_title('RANKING DETALLADO - TOP 15 VARIABLES', fontweight='bold', fontsize=14, pad=20)
                    
                    pdf.savefig(fig_tabla, bbox_inches='tight')
                    plt.close(fig_tabla)
                
                # Leer el PDF generado
                with open(tmp_file.name, 'rb') as f:
                    pdf_data = f.read()
            
            st.success("✅ ¡Informe PDF generado exitosamente!")
            st.info("📄 El informe contiene 7 páginas con análisis completo y gráficos.")
            
            # Botón de descarga del PDF
            st.download_button(
                label="📥 DESCARGAR INFORME COMPLETO (PDF)",
                data=pdf_data,
                file_name=f"informe_micmac_completo_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                mime="application/pdf",
                type="primary"
            )
            
            # Vista previa del contenido
            with st.expander("👁️ Vista Previa del Contenido del Informe PDF"):
                st.markdown("""
                **Página 1:** Portada con resumen ejecutivo  
                **Página 2:** Mapa MICMAC Total con clasificación por cuadrantes  
                **Página 3:** Gráfico del Eje de Estrategia  
                **Página 4:** Ranking de variables (gráfico de barras horizontal)  
                **Página 5:** Grafo de influencias indirectas (si aplica)  
                **Página 6:** Análisis textual detallado  
                **Página 7:** Tabla con ranking detallado top 15 variables  
                
                ✅ **Incluye:** Todos los gráficos generados en la sesión  
                ✅ **Formato:** PDF profesional listo para presentación  
                ✅ **Contenido:** Análisis completo con recomendaciones estratégicas  
                """)


# ============================================================
# DESCARGA EXCEL
# ============================================================
st.markdown("---")
st.markdown("### 💾 Descargar Resultados")

output = io.BytesIO()
with pd.ExcelWriter(output, engine='openpyxl') as writer:
    df_rank.to_excel(writer, sheet_name='Ranking', index=False)
    df_all.to_excel(writer, sheet_name='Completo', index=True)
    
    df_params = pd.DataFrame({
        'Parámetro': ['alpha', 'K', 'Método', 'Fecha', 'Variables'],
        'Valor': [alpha, K_max, 'Mediana' if usar_mediana else 'Media', 
                  datetime.now().strftime("%Y-%m-%d"), len(nombres)]
    })
    df_params.to_excel(writer, sheet_name='Parámetros', index=False)

output.seek(0)

st.download_button(
    "📥 Descargar Excel",
    output,
    f"micmac_a{alpha}_k{K_max}_{datetime.now().strftime('%Y%m%d')}.xlsx",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary"
)

# ============================================================
# FOOTER
# ============================================================
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #666;'>
    <p><strong>MICMAC Interactivo v3.5</strong></p>
    <p>Martín Pratto • 2025</p>
    <p><em>Metodología: Michel Godet (1990)</em></p>
</div>
""", unsafe_allow_html=True)

# SIDEBAR
with st.sidebar:
    st.markdown("### 📖 Guía")
    
    with st.expander("¿Qué es MICMAC?"):
        st.markdown("""
        Método de análisis estructural para identificar variables clave.
        Desarrollado por Michel Godet.
        """)
    
    with st.expander("Interpretación"):
        st.markdown("""
        🔴 **Determinantes:** Palancas de control  
        🔵 **Críticas:** Alta influencia e inestabilidad  
        💧 **Resultado:** Indicadores  
        🟠 **Autónomas:** Independientes  
        """)
